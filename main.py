import streamlit as st
import sqlite3
import os
import uuid
from datetime import datetime
from PIL import Image
import io
import base64
import pandas as pd
APP_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(APP_DIR, "workshop.db")
UPLOAD_DIR = os.path.join(APP_DIR, "uploads")
LOGO_PATH = os.path.join(APP_DIR, "petra_logo.png")
os.makedirs(UPLOAD_DIR, exist_ok=True)
ALARM_THRESHOLD = 3
# ─────────────────────────────────────────────────────────────────────────────
# DATABASE
# ─────────────────────────────────────────────────────────────────────────────
def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn
def init_db():
    conn = get_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS entries_new (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            petra_code TEXT NOT NULL,
            part_number TEXT,
            project_number TEXT,
            notes TEXT,
            image_path TEXT,
            timestamp TEXT NOT NULL
        )
    """)
    existing_tables = [
        r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        ).fetchall()
    ]
    if "entries" in existing_tables and "entries_new" in existing_tables:
        count = conn.execute("SELECT COUNT(*) FROM entries_new").fetchone()[0]
        if count == 0:
            conn.execute("""
                INSERT INTO entries_new
                    (id, petra_code, part_number, project_number, notes, image_path, timestamp)
                SELECT id, petra_code, part_number, project_number, notes, image_path, timestamp
                FROM entries
            """)
        conn.execute("DROP TABLE entries")
        conn.execute("ALTER TABLE entries_new RENAME TO entries")
    elif "entries_new" in existing_tables:
        conn.execute("ALTER TABLE entries_new RENAME TO entries")
    conn.commit()
    conn.close()
def is_duplicate(petra_code, part_number, project_number):
    conn = get_connection()
    pn = project_number.strip() if project_number and project_number.strip() else None
    if pn:
        petra_dup = conn.execute(
            "SELECT COUNT(*) FROM entries WHERE petra_code = ? AND project_number = ?",
            (petra_code, pn)
        ).fetchone()[0]
        if petra_dup > 0:
            conn.close()
            return True, "Petra Code '" + petra_code + "' already exists for Project '" + pn + "'."
        if part_number and part_number.strip():
            part_dup = conn.execute(
                "SELECT COUNT(*) FROM entries WHERE part_number = ? AND project_number = ?",
                (part_number.strip(), pn)
            ).fetchone()[0]
            if part_dup > 0:
                conn.close()
                return True, "Part Number '" + part_number.strip() + "' already exists for Project '" + pn + "'."
    else:
        petra_dup = conn.execute(
            "SELECT COUNT(*) FROM entries WHERE petra_code = ? AND (project_number IS NULL OR project_number = '')",
            (petra_code,)
        ).fetchone()[0]
        if petra_dup > 0:
            conn.close()
            return True, "Petra Code '" + petra_code + "' already exists with no project number."
        if part_number and part_number.strip():
            part_dup = conn.execute(
                "SELECT COUNT(*) FROM entries WHERE part_number = ? AND (project_number IS NULL OR project_number = '')",
                (part_number.strip(),)
            ).fetchone()[0]
            if part_dup > 0:
                conn.close()
                return True, "Part Number '" + part_number.strip() + "' already exists with no project number."
    conn.close()
    return False, ""
def save_entry(petra_code, part_number, project_number, notes, image_path):
    conn = get_connection()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.execute(
        """INSERT INTO entries
           (petra_code, part_number, project_number, notes, image_path, timestamp)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (
            petra_code,
            part_number if part_number else None,
            project_number if project_number else None,
            notes if notes else None,
            image_path,
            timestamp,
        )
    )
    conn.commit()
    conn.close()
def delete_entries(ids):
    if not ids:
        return
    conn = get_connection()
    placeholders = ",".join("?" * len(ids))
    conn.execute("DELETE FROM entries WHERE id IN (" + placeholders + ")", ids)
    conn.commit()
    conn.close()
 def count_after_save(petra_code, part_number):
    conn = get_connection()
    petra_count = conn.execute(
        "SELECT COUNT(*) FROM entries WHERE petra_code = ?", (petra_code,)
    ).fetchone()[0]
    part_count = 0
    if part_number:
        part_count = conn.execute(
            "SELECT COUNT(*) FROM entries WHERE part_number = ?", (part_number,)
        ).fetchone()[0]
    conn.close()
    return petra_count, part_count
def get_all_entries():
    conn = get_connection()
    rows = conn.execute("SELECT * FROM entries ORDER BY timestamp DESC").fetchall()
    conn.close()
    return rows
def get_critical_petra_codes():
    conn = get_connection()
    rows = conn.execute(
        "SELECT petra_code, COUNT(*) AS total_count, MAX(timestamp) AS last_seen "
        "FROM entries GROUP BY petra_code "
        "HAVING COUNT(*) >= " + str(ALARM_THRESHOLD) + " "
        "ORDER BY total_count DESC"
    ).fetchall()
    conn.close()
    return rows
def get_critical_part_numbers():
    conn = get_connection()
    rows = conn.execute(
        "SELECT part_number, COUNT(*) AS total_count, MAX(timestamp) AS last_seen "
        "FROM entries WHERE part_number IS NOT NULL AND part_number != '' "
        "GROUP BY part_number "
        "HAVING COUNT(*) >= " + str(ALARM_THRESHOLD) + " "
        "ORDER BY total_count DESC"
    ).fetchall()
    conn.close()
    return rows
def get_recurring_entries_full():
    conn = get_connection()
    threshold = str(ALARM_THRESHOLD)
    rows = conn.execute(
        "SELECT id, project_number, petra_code, part_number, notes, timestamp "
        "FROM entries "
        "WHERE petra_code IN ("
        "    SELECT petra_code FROM entries "
        "    GROUP BY petra_code HAVING COUNT(*) >= " + threshold
        + ") "
        "OR ("
        "    part_number IS NOT NULL AND part_number != '' AND "
        "    part_number IN ("
        "        SELECT part_number FROM entries "
        "        WHERE part_number IS NOT NULL AND part_number != '' "
        "        GROUP BY part_number HAVING COUNT(*) >= " + threshold
        + "    )"
        ") "
        "ORDER BY petra_code, part_number, timestamp"
    ).fetchall()
    conn.close()
    return rows
def build_excel_report():
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    critical_petra = get_critical_petra_codes()
    critical_parts = get_critical_part_numbers()
    full_rows = get_recurring_entries_full()
    red_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    white_bold = Font(color="FFFFFF", bold=True, size=12)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    def style_sheet(ws):
        for cell in ws[1]:
            cell.fill = red_fill
            cell.font = white_bold
            cell.alignment = center
            cell.border = thin
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 6
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        detail_data = [
            {
                "Project Number": r["project_number"] or "-",
                "Petra Code": r["petra_code"],
                "Part Number": r["part_number"] or "-",
                "Notes": r["notes"] or "-",
                "Submission Date": r["timestamp"],
            }
            for r in full_rows
        ]
        detail_df = pd.DataFrame(detail_data) if detail_data else pd.DataFrame(
            columns=["Project Number", "Petra Code", "Part Number", "Notes", "Submission Date"]
        )
        detail_df.to_excel(writer, index=False, sheet_name="Recurring Entries (Detail)")
        style_sheet(writer.sheets["Recurring Entries (Detail)"])
        petra_rows = [
            {
                "Petra Code": r["petra_code"],
                "Total Occurrences": r["total_count"],
                "Last Reported Date": r["last_seen"],
            }
            for r in critical_petra
        ]
        petra_df = pd.DataFrame(petra_rows) if petra_rows else pd.DataFrame(
            columns=["Petra Code", "Total Occurrences", "Last Reported Date"]
        )
        petra_df.to_excel(writer, index=False, sheet_name="Critical Petra Codes")
        style_sheet(writer.sheets["Critical Petra Codes"])
        part_rows = [
            {
                "Part Number": r["part_number"],
                "Total Occurrences": r["total_count"],
                "Last Reported Date": r["last_seen"],
            }
            for r in critical_parts
        ]
        parts_df = pd.DataFrame(part_rows) if part_rows else pd.DataFrame(
            columns=["Part Number", "Total Occurrences", "Last Reported Date"]
        )
        parts_df.to_excel(writer, index=False, sheet_name="Critical Part Numbers")
        style_sheet(writer.sheets["Critical Part Numbers"])
    output.seek(0)
    return output
def save_image(image_file):
    filename = uuid.uuid4().hex + ".png"
    filepath = os.path.join(UPLOAD_DIR, filename)
    img = Image.open(image_file)
    img.save(filepath)
    return filepath
# ─────────────────────────────────────────────────────────────────────────────
# APP
# ─────────────────────────────────────────────────────────────────────────────
init_db()
st.set_page_config(
    page_title="Panel Workshop - Fault Reporter",
    page_icon="🏭",
    layout="wide"
)
col_l, col_c, col_r = st.columns([1, 2, 1])
with col_c:
    if os.path.exists(LOGO_PATH):
        st.image(LOGO_PATH, width=250)
st.title("Panel Workshop - Fault Reporter")
tab_submit, tab_dashboard, tab_admin = st.tabs(
    ["Submit Entry", "Dashboard", "Admin / Delete"]
)
# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 - SUBMIT ENTRY
# ─────────────────────────────────────────────────────────────────────────────
with tab_submit:
    st.header("Report a Faulty Part")
    st.info("Fields marked with * are required. All other fields are optional. The submission date and time are recorded automatically.")
    with st.form("entry_form", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            petra_code = st.text_input(
                "Petra Code *",
                placeholder="Enter Petra Code (required)",
                help="The Petra code for the faulty part - this is the only required field"
            )
            part_number = st.text_input(
                "Part Number",
                placeholder="Enter Part Number (optional)",
            )
        with col2:
            project_number = st.text_input(
                "Project Number",
                placeholder="Enter Project Number (optional)",
            )
        notes = st.text_area(
            "Notes (optional)",
            placeholder="Describe the fault, observations, or any additional details...",
            height=120
        )
capture_method = st.radio(
            "Choose capture method",
            options=[
                "Use device camera / gallery (recommended for mobile - allows front/back camera switching)",
                "Use in-browser camera (desktop default)",
            ],
            index=0,
            help="On mobile, the device camera option lets you choose front camera, back camera, or pick from your gallery."
        )
        camera_image = None
        if capture_method.startswith("Use device"):
            camera_image = st.file_uploader(
                "Tap to open your camera or pick a photo",
                type=["png", "jpg", "jpeg"],
                accept_multiple_files=False,
                help="On mobile this opens your device's native picker - choose front camera, back camera, or an existing photo."
            )
        else:
            camera_image = st.camera_input("Take a photo of the faulty part")
        submitted = st.form_submit_button(
            "Submit Report", use_container_width=True, type="primary"
        )
        if submitted:
            if not petra_code.strip():
                st.error("Petra Code is required. Please enter a Petra Code before submitting.")
            else:
                dup, dup_reason = is_duplicate(
                    petra_code.strip(), part_number.strip(), project_number.strip()
                )
                if dup:
                    st.error("Duplicate entry blocked: " + dup_reason + " This combination has already been submitted. Please verify the data before resubmitting.")
                else:
                    image_path = None
                    if camera_image:
                        image_path = save_image(camera_image)
                    save_entry(
                        petra_code.strip(),
                        part_number.strip(),
                        project_number.strip(),
                        notes.strip(),
                        image_path,
                    )
                    petra_count, part_count = count_after_save(
                        petra_code.strip(), part_number.strip()
                    )
                    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    st.success("Entry submitted successfully! Recorded at: " + now_str)
                    alarm_petra = petra_count >= ALARM_THRESHOLD
                    alarm_part = part_number.strip() and part_count >= ALARM_THRESHOLD
                    if alarm_petra or alarm_part:
                        st.error("HIGH PRIORITY - This part has repeated issues. Please check EPLAN routing/definition.")
                        if alarm_petra:
                            st.warning("Petra Code '" + petra_code + "' has now been reported " + str(petra_count) + " time(s).")
                        if alarm_part:
                            st.warning("Part Number '" + part_number + "' has now been reported " + str(part_count) + " time(s).")
# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 - DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
with tab_dashboard:
    critical_petra = get_critical_petra_codes()
    critical_parts = get_critical_part_numbers()
    has_critical = len(critical_petra) > 0 or len(critical_parts) > 0
    st.header("Critical Recurring Issues")
    st.caption("Items reported " + str(ALARM_THRESHOLD) + " or more times.")
    if not has_critical:
        st.success("No recurring issues detected. All parts are within acceptable submission limits.")
    else:
        st.error("The following codes have been reported " + str(ALARM_THRESHOLD) + "+ times and require immediate EPLAN review.")
        col_petra, col_parts = st.columns(2)
        with col_petra:
            st.subheader("Recurring Petra Codes")
            if critical_petra:
                st.dataframe(
                    pd.DataFrame([{
                        "Petra Code": r["petra_code"],
                        "Occurrences": r["total_count"],
                        "Last Reported": r["last_seen"],
                    } for r in critical_petra]),
                    use_container_width=True, hide_index=True,
                )
            else:
                st.info("None found.")
        with col_parts:
            st.subheader("Recurring Part Numbers")
            if critical_parts:
                st.dataframe(
                    pd.DataFrame([{
                        "Part Number": r["part_number"],
                        "Occurrences": r["total_count"],
                        "Last Reported": r["last_seen"],
                    } for r in critical_parts]),
                    use_container_width=True, hide_index=True,
                )
            else:
                st.info("None found.")
    st.markdown("---")
    st.header("Export Critical Issues")
    st.info("Generates an Excel file with 3 sheets: full detail rows for all recurring entries (Project Number, Petra Code, Part Number, Notes, Submission Date), plus summary sheets for critical codes and part numbers.")

    if has_critical:
        excel_data = build_excel_report()
        st.download_button(
            label="Download Critical Issues Report (Excel)",
            data=excel_data,
            file_name="Critical_Issues_Report_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
        total_critical = len(critical_petra) + len(critical_parts)
        st.caption(
            "Report covers " + str(len(critical_petra)) + " Petra Code(s) and "
            + str(len(critical_parts)) + " Part Number(s) - " + str(total_critical) + " critical item(s) total."
        )
    else:
        st.info("No critical issues to export yet.")

    st.markdown("---")
    st.header("All Submitted Entries")

    entries = get_all_entries()

    if not entries:
        st.info("No entries yet. Submit your first fault report using the 'Submit Entry' tab.")
    else:
        st.write("Total entries: " + str(len(entries)))
        search_term = st.text_input(
            "Search by Petra Code, Part Number, or Project Number", ""
        )

        filtered = entries
        if search_term:
            s = search_term.lower()
            filtered = [
                e for e in entries
                if s in (e["petra_code"] or "").lower()
                or s in (e["part_number"] or "").lower()
                or s in (e["project_number"] or "").lower()
            ]

        st.write("Showing " + str(len(filtered)) + " record(s)")

        for entry in filtered:
            label = str(entry["timestamp"]) + " | Petra: " + str(entry["petra_code"])
            if entry["part_number"]:
                label += " | Part: " + str(entry["part_number"])
            if entry["project_number"]:
                label += " | Project: " + str(entry["project_number"])

            with st.expander(label, expanded=False):
                col_info, col_img = st.columns([2, 1])

                with col_info:
                    st.write("Submission Date: " + str(entry["timestamp"]))
                    st.write("Petra Code: " + str(entry["petra_code"]))
                    st.write("Part Number: " + (str(entry["part_number"]) if entry["part_number"] else "-"))
                    st.write("Project Number: " + (str(entry["project_number"]) if entry["project_number"] else "-"))
                    st.write("Notes: " + (str(entry["notes"]) if entry["notes"] else "No notes provided"))

                with col_img:
                    if entry["image_path"] and os.path.exists(entry["image_path"]):
                        st.image(
                            entry["image_path"], caption="Captured Photo",
                            use_container_width=True
                        )
                    else:
                        st.write("No image captured")


# ─────────────────────────────────────────────────────────────────────────────
# TAB 3 - ADMIN / DELETE
# ─────────────────────────────────────────────────────────────────────────────
with tab_admin:
    st.header("Admin - Delete Entries")
    st.warning("Deleted entries cannot be recovered. Use this section only to remove accidental or test submissions.")

    all_entries = get_all_entries()

    if not all_entries:
        st.info("No entries in the database.")
    else:
        entry_options = {}
        for e in all_entries:
            label = "[ID " + str(e["id"]) + "] " + str(e["timestamp"]) + " - Petra: " + str(e["petra_code"])
            if e["part_number"]:
                label += " | Part: " + str(e["part_number"])
            if e["project_number"]:
                label += " | Project: " + str(e["project_number"])
            entry_options[label] = e["id"]

        selected_labels = st.multiselect(
            "Select entries to delete",
            options=list(entry_options.keys()),
            help="You can select multiple entries at once."
        )

        if selected_labels:
            st.write(str(len(selected_labels)) + " entry(ies) selected for deletion.")

            for label in selected_labels:
                eid = entry_options[label]
                entry = next((e for e in all_entries if e["id"] == eid), None)
                if entry:
                    with st.expander("Preview: " + label, expanded=False):
                        st.write("Submission Date: " + str(entry["timestamp"]))
                        st.write("Petra Code: " + str(entry["petra_code"]))
                        st.write("Part Number: " + (str(entry["part_number"]) if entry["part_number"] else "-"))
                        st.write("Project Number: " + (str(entry["project_number"]) if entry["project_number"] else "-"))
                        st.write("Notes: " + (str(entry["notes"]) if entry["notes"] else "None"))

            confirm = st.checkbox(
                "I confirm I want to permanently delete " + str(len(selected_labels)) + " entry(ies)."
            )

            if st.button(
                "Delete " + str(len(selected_labels)) + " Selected Entry(ies)",
                type="primary",
                disabled=not confirm,
            ):
                ids_to_delete = [entry_options[lbl] for lbl in selected_labels]
                delete_entries(ids_to_delete)
                st.success(str(len(ids_to_delete)) + " entry(ies) deleted successfully. Refresh the page to see updated data.")
                st.rerun()
        else:
            st.info("Select one or more entries above to delete them.")

        st.subheader("Capture Image (optional)")
        
